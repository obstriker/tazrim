from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pickle
from flask import Flask, jsonify, request
from flask import Flask, request, send_file

app = Flask(__name__)


class ExcelPopulator:
    def __init__(self, filename):
        self.filename = filename
        self.workbook = load_workbook(filename=filename)
        self.sheet = self.workbook.active

    def populate_table(self, table, start_row=1, start_col=1):
        for row_index, row in enumerate(table.get_data()):
            for col_index, value in enumerate(row):
                self.sheet.cell(row=start_row + row_index, column=start_col + col_index, value=value)
    
    def add_data(self, data, start_row=1, start_col=1):
        for row_index, row in enumerate(data):
            for col_index, key in enumerate(row.keys()):

                # For loop a nested object
                if key == "monthly_table":
                    # Process monthly_table separately
                    monthly_table = row[key]
                    for month_index, (month, value) in enumerate(monthly_table.items()):
                        self.sheet.cell(row=start_row + row_index, column=start_col + col_index + month_index * 2, value=value)
                        self.sheet.cell(row=start_row + row_index, column=start_col + col_index + month_index * 2 - 1, value="")
                    continue  # Skip the rest of the loop for this row

                self.sheet.cell(row=start_row + row_index, column=start_col + col_index, value=row[key])

                # print(col_index, value)
                
            # for loop on monthly_table only, takes 2 cells per iteration
            # continues from where col_index stopped



            # Fill the header row with color
            # if row_index == 0:  # Assuming the first row is the header
                # self.fill_header(start_row)

    def fill_header(self, row_index):
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
        for cell in self.sheet[row_index]:  # Fill the entire header row
            cell.fill = fill

    def save(self, filename="file.xlsx"):
        self.workbook.save(filename)


# {income:[], expenses:[], product_expenses:[], project_payments:[], media_payments:[], media_one_time_payment:[]}

@app.route('/generate', methods=['POST'])
def generate_excel():
    new_data = request.json
    excel_populator = ExcelPopulator("template_test.xlsx")

    excel_populator.add_data(data=new_data["income"], start_row=3, start_col=1)
    excel_populator.add_data(data=new_data["expenses"], start_row=16, start_col=1)
    excel_populator.add_data(data=new_data["product_expenses"], start_row=27, start_col=1)
    excel_populator.add_data(data=new_data["project_payments"], start_row=39, start_col=1)
    excel_populator.add_data(data=new_data["media_payments"], start_row=44, start_col=1)
    excel_populator.add_data(data=new_data["media_one_time_payment"], start_row=61, start_col=1)
    # income_data.append(new_data)

    # Save to an in-memory BytesIO stream
    output = BytesIO()
    excel_populator.save(output)
    output.seek(0)  # Move to the beginning of the stream

    # Send the Excel file as a download response
    return send_file(output, as_attachment=True, download_name="annual_cashflow_report.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


data_example = {
    "income": [
        {
            "source_name": "Salary",
            "total_income": 84000,
            "last_year": 80000,
            "predicted_yearly_income": 85000,
            "expenses": 30000,
            "monthly_table": {
                "January": 7000,
                "February": 7000,
                "March": 7000,
                "April": 7000,
                "May": 7000,
                "June": 7000,
                "July": 7000,
                "August": 7000,
                "September": 7000,
                "October": 7000,
                "November": 7000,
                "December": 7000
            },
        },
        {
            "source_name": "Freelance",
            "total_income": 30000,
            "monthly_table": {
                "January": 2500,
                "February": 2500,
                "March": 2500,
                "April": 2500,
                "May": 2500,
                "June": 2500,
                "July": 2500,
                "August": 2500,
                "September": 2500,
                "October": 2500,
                "November": 2500,
                "December": 2500
            },
            "last_year": 28000,
            "predicted_yearly_income": 32000,
            "expenses": 10000
        },
        {
            "source_name": "Investments",
            "total_income": 6000,
            "monthly_table": {
                "January": 500,
                "February": 500,
                "March": 500,
                "April": 500,
                "May": 500,
                "June": 500,
                "July": 500,
                "August": 500,
                "September": 500,
                "October": 500,
                "November": 500,
                "December": 500
            },
            "last_year": 5000,
            "predicted_yearly_income": 7000,
            "expenses": 2000
        },
        {
            "source_name": "Consulting",
            "total_income": 48000,
            "last_year": 45000,
            "predicted_yearly_income": 50000,
            "expenses": 15000,
            "monthly_table": {
                "January": 4000,
                "February": 4000,
                "March": 4000,
                "April": 4000,
                "May": 4000,
                "June": 4000,
                "July": 4000,
                "August": 4000,
                "September": 4000,
                "October": 4000,
                "November": 4000,
                "December": 4000
            },
        },
        {
            "source_name": "Rental Income",
            "total_income": 18000,
            "last_year": 17000,
            "predicted_yearly_income": 19000,
            "expenses": 6000,
            "monthly_table": {
                "January": 1500,
                "February": 1500,
                "March": 1500,
                "April": 1500,
                "May": 1500,
                "June": 1500,
                "July": 1500,
                "August": 1500,
                "September": 1500,
                "October": 1500,
                "November": 1500,
                "December": 1500
            },
        }
    ],
    "expenses": [
        {
            "category_name": "Rent",
            "total_expense": 18000,
            "monthly_table": {
                "January": 1500,
                "February": 1500,
                "March": 1500,
                "April": 1500,
                "May": 1500,
                "June": 1500,
                "July": 1500,
                "August": 1500,
                "September": 1500,
                "October": 1500,
                "November": 1500,
                "December": 1500
            }
        },
        {
            "category_name": "Utilities",
            "total_expense": 4800,
            "monthly_table": {
                "January": 400,
                "February": 400,
                "March": 400,
                "April": 400,
                "May": 400,
                "June": 400,
                "July": 400,
                "August": 400,
                "September": 400,
                "October": 400,
                "November": 400,
                "December": 400
            }
        },
        {
            "category_name": "Groceries",
            "total_expense": 6000,
            "monthly_table": {
                "January": 500,
                "February": 500,
                "March": 500,
                "April": 500,
                "May": 500,
                "June": 500,
                "July": 500,
                "August": 500,
                "September": 500,
                "October": 500,
                "November": 500,
                "December": 500
            }
        },
        {
            "category_name": "Transportation",
            "total_expense": 2400,
            "monthly_table": {
                "January": 200,
                "February": 200,
                "March": 200,
                "April": 200,
                "May": 200,
                "June": 200,
                "July": 200,
                "August": 200,
                "September": 200,
                "October": 200,
                "November": 200,
                "December": 200
            }
        },
        {
            "category_name": "Insurance",
            "total_expense": 3600,
            "monthly_table": {
                "January": 300,
                "February": 300,
                "March": 300,
                "April": 300,
                "May": 300,
                "June": 300,
                "July": 300,
                "August": 300,
                "September": 300,
                "October": 300,
                "November": 300,
                "December": 300
            }
        }
    ],
    "product_expenses": [
        {
            "product_name": "Product A",
            "total_expense": 1000,
            "monthly_table": {
                "January": 100,
                "February": 100,
                "March": 100,
                "April": 100,
                "May": 100,
                "June": 100,
                "July": 100,
                "August": 100,
                "September": 100,
                "October": 100,
                "November": 100,
                "December": 100
            }
        },
        {
            "product_name": "Product B",
            "total_expense": 2400,
            "monthly_table": {
                "January": 200,
                "February": 200,
                "March": 200,
                "April": 200,
                "May": 200,
                "June": 200,
                "July": 200,
                "August": 200,
                "September": 200,
                "October": 200,
                "November": 200,
                "December": 200
            }
        },
        {
            "product_name": "Product C",
            "total_expense": 1800,
            "monthly_table": {
                "January": 150,
                "February": 150,
                "March": 150,
                "April": 150,
                "May": 150,
                "June": 150,
                "July": 150,
                "August": 150,
                "September": 150,
                "October": 150,
                "November": 150,
                "December": 150
            }
        },
        {
            "product_name": "Product D",
            "total_expense": 3600,
            "monthly_table": {
                "January": 300,
                "February": 300,
                "March": 300,
                "April": 300,
                "May": 300,
                "June": 300,
                "July": 300,
                "August": 300,
                "September": 300,
                "October": 300,
                "November": 300,
                "December": 300
            }
        }
    ],
    "project_payments": [
        {
            "project_name": "Project X",
            "total_payment": 18000,
            "monthly_table": {
                "January": 1500,
                "February": 1600,
                "March": 1700,
                "April": 1800,
                "May": 1900,
                "June": 2000,
                "July": 2100,
                "August": 2200,
                "September": 2300,
                "October": 2400,
                "November": 2500,
                "December": 2600
            }
        },
        {
            "project_name": "Project Y",
            "total_payment": 30000,
            "monthly_table": {
                "January": 2500,
                "February": 2600,
                "March": 2700,
                "April": 2800,
                "May": 2900,
                "June": 3000,
                "July": 3100,
                "August": 3200,
                "September": 3300,
                "October": 3400,
                "November": 3500,
                "December": 3600
            }
        },
        {
            "project_name": "Project Z",
            "total_payment": 36000,
            "monthly_table": {
                "January": 3000,
                "February": 3100,
                "March": 3200,
                "April": 3300,
                "May": 3400,
                "June": 3500,
                "July": 3600,
                "August": 3700,
                "September": 3800,
                "October": 3900,
                "November": 4000,
                "December": 4100
            }
        }
    ],
    "media_payments": [
        {
            "campaign_name": "Ad Campaign",
            "total_payment": 8000,
            "monthly_table": {
                "January": 800,
                "February": 800,
                "March": 800,
                "April": 800,
                "May": 800,
                "June": 800,
                "July": 800,
                "August": 800,
                "September": 800,
                "October": 800,
                "November": 800,
                "December": 800
            }
        },
        {
            "campaign_name": "Social Media",
            "total_payment": 3600,
            "monthly_table": {
                "January": 300,
                "February": 320,
                "March": 340,
                "April": 360,
                "May": 380,
                "June": 400,
                "July": 420,
                "August": 440,
                "September": 460,
                "October": 480,
                "November": 500,
                "December": 520
            }
        },
        {
            "campaign_name": "Email Marketing",
            "total_payment": 6000,
            "monthly_table": {
                "January": 500,
                "February": 520,
                "March": 540,
                "April": 560,
                "May": 580,
                "June": 600,
                "July": 620,
                "August": 640,
                "September": 660,
                "October": 680,
                "November": 700,
                "December": 720
            }
        }
    ],
    "media_one_time_payment": [
        {
            "service_name": "Website Development",
            "last_year": 2000,
            "predicted_yearly_income": 2000,
            "expenses": 3000,
            "monthly_table": {
                "January": 500,
                "February": 520,
                "March": 540,
                "April": 560,
                "May": 580,
                "June": 600,
                "July": 620,
                "August": 640,
                "September": 660,
                "October": 680,
                "November": 700,
                "December": 720
            }
        },
        {
            "service_name": "Website Development",
            "last_year": 2000,
            "predicted_yearly_income": 2000,
            "expenses": 3000,
            "monthly_table": {
                "January": 500,
                "February": 520,
                "March": 540,
                "April": 560,
                "May": 580,
                "June": 600,
                "July": 620,
                "August": 640,
                "September": 660,
                "October": 680,
                "November": 700,
                "December": 720
            }
        }
    ]
}

if __name__ == "__main__":
    # Create an instance of ExcelPopulator
    excel_populator = ExcelPopulator("template_test.xlsx")

    # Add headers to the sheet
    # excel_populator.add_data([headers], start_row=1, start_col=1)  # Assuming you want headers in row 1

    # Add data to the sheet starting from row 2
    # excel_populator.add_data(income_data, start_row=2, start_col=1)
    # Create an instance of incomeTable
    # income_table = incomeTable(data=income_data)

    # Insert a new row
    # income_table.insert_row("Alice Johnson", 80000, 90000, 85000, [8000, 9000, 10000])
    # income_table.install_table(excel_populator.sheet, 4,1)
    excel_populator.add_data(data=data_example["income"], start_row=3, start_col=1)
    excel_populator.add_data(data=data_example["expenses"], start_row=16, start_col=1)
    excel_populator.add_data(data=data_example["product_expenses"], start_row=27, start_col=1)
    excel_populator.add_data(data=data_example["project_payments"], start_row=39, start_col=1)
    excel_populator.add_data(data=data_example["media_payments"], start_row=44, start_col=1)
    excel_populator.add_data(data=data_example["media_one_time_payment"], start_row=61, start_col=1)
    # income_data.append(data_example)

    # Save the workbook after installation
    excel_populator.save()


    if __name__ == "__main__":
        app.run(debug=True)