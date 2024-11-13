from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
import pickle
from flask import Flask, jsonify, request

app = Flask(__name__)

# Create class for each table
# Each table class has styling headers, populate data, save to worksheet
#       - Handling and managing table object
# Create main that generate the whole excel file (multiple tables with dynamic sizes)
# Each table returns how much rows it took
# start with one table: income - name, last_year,future,today, income_by_month[]

# Add rest api

# What about excel viewing later on with react? 

# Header is styled header
# Footer is footer with all the formulas
# incomeTablestyles = {'headers': '', 'footer': ''}

# Prepare pkl data as following:
# pkl["headers"] = {"income": ["header": headers, "footer": footer...],
#                    "expenses": ["header": headers, "footer": footer...], "product_expenses": [], "project_payments": [], "media_payments": [], "media_one_time_payment": []}

# prepared pkls
# Create class to install tables and populate data

def represents_int(s):
    try: 
        int(s)
    except ValueError:
        return False
    else:
        return True

def copy_cell(source_cell, target_cell):
    # Copy the value
    target_cell.value = source_cell.value
    
    # Copy the style properties
    target_cell.font = source_cell.font.copy()
    target_cell.fill = source_cell.fill.copy()
    target_cell.alignment = source_cell.alignment.copy()
    target_cell.border = source_cell.border.copy()
    target_cell.number_format = source_cell.number_format
    target_cell.data_type = source_cell.data_type

    # Copy additional properties
    target_cell.hyperlink = source_cell.hyperlink  # Copy hyperlink
    if source_cell.comment:
        target_cell.comment = source_cell.comment.copy()  # Copy comment
    # if source_cell.data_validation:
        # target_cell.data_validation = source_cell.data_validation.copy()  # Copy data validation
    target_cell.protection = source_cell.protection.copy()  # Copy protection settings

def copy_worksheet_properties(source_worksheet, target_worksheet):
    # Copy row heights
    for row in source_worksheet.row_dimensions:
        target_worksheet.row_dimensions[row].height = source_worksheet.row_dimensions[row].height

    # Copy column widths
    for col in source_worksheet.column_dimensions:
        target_worksheet.column_dimensions[col].width = source_worksheet.column_dimensions[col].width

    # Copy sheet view properties (including direction)
    target_worksheet.sheet_view.rightToLeft = source_worksheet.sheet_view.rightToLeft

class incomeTable:
    def __init__(self, data, row = 2, col = 2):
        self.start_row = row
        self.start_col = col
        self.data = data

        # Load the object from the file
        with open('tables.pkl', 'rb') as f:
            pkl = pickle.load(f)
        
        self.headers = pkl["expenses"]["headers"]
        self.footer = pkl["expenses"]["footer"][0]
        
    def get_rows(self):
        # return number of rows including headers according to data entries
        pass

    def insert_row(self, name, last_year,future,today, income_by_month):
        row = [name, last_year,future,today]
        for i in income_by_month:
            row.append(i)
        self.data.append(row)

    def get_data(self):
        return [self.headers] + self.data  # Include headers in the output

    def style_table(self):
        
        pass

    def install_table(self, worksheet, row, col):
        # Set headers
        for j, header in enumerate(self.headers[0]):
            c = worksheet.cell(row=row, column=col + j, value=header.value)  # Set header values
            copy_cell(header, c)
            # Add border to header cell
            # c.border = header.border.copy()  # Copy border style

        for j, header in enumerate(self.headers[1]):
            c = worksheet.cell(row=row + 1, column=col + j, value=header.value)  # Set header values
            copy_cell(header, c)

            # Add border to header cell
            # c.border = header.border.copy()  # Copy border style

        # Set data
        for i, data_row in enumerate(self.data):
            for j, value in enumerate(data_row):
                cell = worksheet.cell(row=row + 2 + i, column=col + j, value=value)  # Set data values

                # Add border to data cell
                cell.border = header.border.copy()  # Copy border style

            # Apply font size and background color to the entire row
            for j in range(len(data_row)):
                cell = worksheet.cell(row=row + 2 + i, column=col + j)
                cell.font = cell.font.copy(size=14)  # Set font size to 14 (or any desired size)
            
                # Alternate background color for the entire row
                if i % 2 == 0:
                    cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")  # Light green
                else:
                    cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # White

                # Check if the value is a number and apply currency format
                if isinstance(cell.value, (int, float)) or (isinstance(cell.value, (str)) and represents_int(cell.value)):
                    cell.number_format = '"₪"#,##0.00'  # Set number format to ILS currency

        # Install footer
        footer_row = row + 2 + len(self.data)  # Calculate the row for the footer
        for j, footer_value in enumerate(self.footer):
            # Assuming the footer is a SUM formula, adjust the range accordingly
            sum_formula = f"=SUM({worksheet.cell(row=row + 2, column=col + j).coordinate}:{worksheet.cell(row=footer_row - 1, column=col + j).coordinate})"
            c = worksheet.cell(row=footer_row, column=col + j, value=sum_formula)  # Set footer values with SUM formula
            copy_cell(footer_value, c)  # Copy footer styles

    def save(self, file):
        pass

class ExcelPopulator:
    def __init__(self, filename=""):
        if filename == "":
            self.workbook = Workbook()
            self.filename = "template.xlsx"
        else:
            self.workbook = load_workbook(filename=filename)
        self.sheet = self.workbook.active

    def populate_table(self, table, start_row=1, start_col=1):
        for row_index, row in enumerate(table.get_data()):
            for col_index, value in enumerate(row):
                self.sheet.cell(row=start_row + row_index + 2, column=start_col + col_index, value=value)
    
    def add_data(self, data, start_row=1, start_col=1):
        for row_index, row in enumerate(data):
            for col_index, value in enumerate(row):
                try:
                    self.sheet.cell(row=start_row + row_index + 2, column=start_col + col_index, value=value)
                except:
                    continue
                # print(col_index, value)
            # Fill the header row with color
            # if row_index == 0:  # Assuming the first row is the header
                # self.fill_header(start_row)

    def fill_header(self, row_index):
        fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # Yellow fill
        for cell in self.sheet[row_index]:  # Fill the entire header row
            cell.fill = fill

    def save(self, filename="template.xlsx"):
        self.workbook.save(filename)


# {income:[], expenses:[], product_expenses:[], project_payments:[], media_payments:[], media_one_time_payment:[]}

@app.route('/generate', methods=['POST'])
def generate_excel():
    new_data = request.json
    income_data.append(new_data)
    return jsonify(new_data), 201


def create_pkl():
    pkl = {"income": {"headers": [], "footer": []},
            "expenses": {"headers": [], "footer": []},
            "product_expenses": {"headers": [], "footer": []},
            "project_payments": {"headers": [], "footer": []}, 
            "media_payments": {"headers": [], "footer": []}, 
            "media_one_time_payment": {"headers": [], "footer": []},
            "summary": {"headers": [], "footer": []}
            }
    

    x = ExcelPopulator("tazrim.xlsx").sheet

    for i, row in enumerate(x.rows):

        # income table
        if i in [0,1]:
            pkl["income"]["headers"].append(row)
            print([i.value for i in row])
        elif i in [12]:
            pkl["income"]["footer"].append(row)
        
        # expense table
        if i in [13,14]:
            pkl["expenses"]["headers"].append(row)
        elif i in [24]:
            pkl["expenses"]["footer"].append(row)

        #  product_expenses table
        if i in [25]:
            pkl["product_expenses"]["headers"].append(row)
        elif i in [36]:
            pkl["product_expenses"]["footer"].append(row)

        # project_payments table
        if i in [37]:
            pkl["project_payments"]["headers"].append(row)
        elif i in [41]:
            pkl["project_payments"]["footer"].append(row)

        # media_payments table
        if i in [42]:
            pkl["media_payments"]["headers"].append(row)
        elif i in [58]:
            pkl["media_payments"]["footer"].append(row)

        # media_one_time_payment table
        if i in [59]:
            pkl["media_one_time_payment"]["headers"].append(row)
        elif i in [62]:
            pkl["media_one_time_payment"]["footer"].append(row)

        if i in [i for i in range(65,75)]:
            pkl["summary"]["headers"].append(row)
            pkl["summary"]["footer"] = []

    f = open("tables.pkl", "wb")
    pickle.dump(pkl, f)
    # pkl["income"]["header"] = [cell cell in x.sheet.ce]

if __name__ == "__main__":
    # Create an instance of ExcelPopulator
    excel_populator = ExcelPopulator()

    # Define the headers and data
    headers = ["Name", "Last Year", "Future", "Today"]
    income_data = [
        ["כסף על הראש שלי", "5000", "10000", "0"],
        ["Dummy Name 1", "1000", "2000", "5",""  ,1,"",2,"",3,"",4,"",5,"",6,"",6,"",7,"",8,"",9,"",10,"",11,"",12],
        ["Dummy Name 2", "1500", "2500", "10", "",1,"",2,"",3,"",4,"",5,"",6,"",6,"",7,"",8,"",9,"",10,"",11,"",12],
        ["Dummy Name 3", "2000", "3000", "15", "",1,"",2,"",3,"",4,"",5,"",6,"",6,"",7,"",8,"",9,"",10,"",11,"",12]
    ]

    # Add headers to the sheet
    # excel_populator.add_data([headers], start_row=1, start_col=1)  # Assuming you want headers in row 1

    template = ExcelPopulator("tazrim.xlsx")
    copy_worksheet_properties(template.workbook.active, excel_populator.workbook.active)
    # Add data to the sheet starting from row 2
    # excel_populator.add_data(income_data, start_row=2, start_col=1)
    # Create an instance of incomeTable
    income_table = incomeTable(data=income_data)
    # create_pkl()
    # Insert a new row
    income_table.insert_row("Alice Johnson", 80000, 90000, 85000, [8000, 9000, 10000])
    income_table.install_table(excel_populator.sheet, 8,1)
    income_table.install_table(excel_populator.sheet, 20,1)

    # Save the workbook after installation
    excel_populator.save("template.xlsx")